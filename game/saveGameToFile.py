import os
import openpyxl

file = 'game_stats.xlsx'
fileExists = os.path.isfile(file)

def openExcelFile():
    if fileExists:
        wb = openpyxl.load_workbook(file)
    else:
        wb = openpyxl.Workbook()
        activeSheet = wb.active
        activeSheet['A1'].value = 'Master BIG DATA'
        activeSheet['A2'].value = 'Juego Adivina el Número (Tarea Final, Python Básico)'
        activeSheet['A3'].value = 'Alberto Patiño'
        activeSheet['A4'].value = ''
        activeSheet['A5'].value = 'Jugadores'
        activeSheet['B5'].value = 'Partidas Jugadas'
        activeSheet['C5'].value = 'Ganadas'
        activeSheet['D5'].value = 'Perdidas'
        activeSheet['E5'].value = 'Porcentaje de aciertos'
    return wb


def saveGameToFile(gameData):
    wb = openExcelFile()
    activeSheet = wb.active
    maxColumn = activeSheet.max_column
    maxRow = activeSheet.max_row
    isPlayerNotRegistered = True

    for row in range(6, maxRow+1):
        if activeSheet.cell(row=row, column=1).value == gameData[0]: # player name
            isPlayerNotRegistered = False                        
            totalGames = activeSheet.cell(row=row, column=2).value # total games
            totalGames + 1
            totalWins = activeSheet.cell(row=row, column=3).value # total wins
            totalLoses = activeSheet.cell(row=row, column=4).value # total loses
            if gameData[1] == 'Win':
               totalWins + 1
            else:
                totalLoses + 1
            winnigPercentage = (totalWins / totalGames) * 100
            activeSheet.cell(row=row, column=5).value = round(winningPercentage) # winnig percentage

            break        
    if isPlayerNotRegistered:
        activeSheet.cell(row=maxRow+1, column=1).value = gameData[0] # player name
        activeSheet.cell(row=maxRow+1, column=2).value = 1 # total games
        if gameData[1] == 'Win':
            activeSheet.cell(row=maxRow+1, column=3).value = 1 # games won
            activeSheet.cell(row=maxRow+1, column=5).value = 100 # winned percentage
        else:
            activeSheet.cell(row=maxRow+1, column=4).value = 1 # games lost
            activeSheet.cell(row=maxRow+1, column=5).value = 0 # winned percentage

    wb.save(file)
