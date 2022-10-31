import os
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

file = 'game_stats.xlsx'
fileExists = os.path.isfile(file)


def showStatistics():
    if fileExists:
        wb = openpyxl.load_workbook(file)
        activeSheet = wb.active
        maxColumn = activeSheet.max_column
        maxRow = activeSheet.max_row

        columns = []
        for column in range(1, maxColumn+1):
            columns.append(activeSheet.cell(row=5, column=column).value)

        data = []
        for row in range(6, maxRow+1):
            rowValues = []
            for column in range(1, maxColumn+1):
                rowValues.append(activeSheet.cell(row=row, column=column).value)
            data.append(rowValues)
            
        df = pd.DataFrame(data, columns=columns)
        
        df.plot(x='Jugadores', 
                kind='bar',
                grid=True,
                rot=45,
                # stacked=True,
                title='Juego Adivina el Número')        
        plt.show()
    else:
        print('No hay estadísticas aún.\n')

